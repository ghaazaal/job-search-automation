"""
main.py — Job Hunt Automation
==============================
Usage:
  python main.py                        # scrape + score + persist + serve (blocks until Ctrl-C)
  python main.py --serve                # serve the existing store without scraping
  python main.py --tailor --jd job.txt  # tailor one job from a saved JD file

Setup:
  PowerShell:  $env:APIFY_TOKEN = "apify_api_..."
               $env:ANTHROPIC_API_KEY = "sk-ant-..."
  pip install -r requirements.txt
"""
import argparse
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import yaml
from dotenv import load_dotenv

# ── Config ────────────────────────────────────────────────────────────────────
_BASE = Path(__file__).resolve().parent
CONFIG_PATH     = _BASE / "config.yaml"
VOCABULARY_PATH = _BASE / "vocabulary.yaml"

# One user, no authentication — see the Identity section of the design spec.
# Schema version 3 renames any pre-existing user to this.
DEFAULT_USER = "default"

# Load .env from the job_hunt/ directory (where this file lives).
# os.environ values already set take precedence — dotenv never overwrites them.
load_dotenv(_BASE / ".env")


def _load_config() -> dict:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ── Imports ───────────────────────────────────────────────────────────────────
_PROVIDER_PKG = {
    "anthropic": "anthropic",
    "groq":      "groq",
    "gemini":    "google.generativeai",
}


def _check_deps(config: dict) -> None:
    """Check required packages are installed. Fails fast with a clear message."""
    provider = (
        os.environ.get("LLM_PROVIDER")
        or config.get("llm", {}).get("provider", "anthropic")
    ).lower()

    always_needed = ("openpyxl", "yaml", "requests", "dotenv")
    provider_pkg  = _PROVIDER_PKG.get(provider, "")

    missing = []
    for pkg in (*always_needed, provider_pkg):
        if not pkg:
            continue
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg.replace("_", "-").replace(".", "-"))
    if missing:
        print(f"ERROR: Missing packages: {', '.join(missing)}")
        print(f"Run: pip install -r {_BASE / 'requirements.txt'}")
        sys.exit(1)


def _open_store():
    """A connection with the schema present. The caller closes it."""
    from src.store.db import connect
    from src.store.ingest import ensure_user
    from src.store.schema import init_db

    conn = connect()
    init_db(conn)
    return conn, ensure_user(conn, DEFAULT_USER)


def _announce_step(step: dict, printed_steps: set[tuple[str, str, str]]) -> None:
    """Print one scrape step's result the first time it's seen done/failed.

    `printed_steps` is owned by the caller and must persist across calls —
    `event["steps"]` handed to `_announce` is a fresh copy every progress
    event (run_core.execute's report() snapshots it so retained events can't
    be mutated later), so de-duplication can't be tracked by mutating the
    step dict itself; it has to live in a set the caller keeps alive for the
    whole run.
    """
    key = (step["source"], step["role"], step.get("lane", "worldwide"))
    label = step["source"]
    lane = step.get("lane", "worldwide")
    if lane != "worldwide":
        label += f" · {lane}"
    if step.get("state") == "done" and key not in printed_steps:
        printed_steps.add(key)
        print(f"  {label} / {step['role']}...")
        print(f"    -> {step['found']} fetched")
    elif step.get("state") == "failed" and key not in printed_steps:
        printed_steps.add(key)
        print(f"  {label} / {step['role']}... FAILED")


logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
                    datefmt="%H:%M:%S")
logger = logging.getLogger("main")

# Fix Windows cp1252 terminal encoding — print non-ASCII chars safely
import sys as _sys
if hasattr(_sys.stdout, "reconfigure"):
    try:
        _sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# ── Pipeline ──────────────────────────────────────────────────────────────────
def run_pipeline(config: dict) -> None:
    from src.scrapers import indeed as indeed_scraper
    from src.scrapers import linkedin as li_scraper
    from src.scoring.scorer import Scorer
    from src.agents.resume_agent import generate_role_variant, _safe
    from src.agents.tailoring_agent import shortlist_decision
    from src.store.profile import get_profile, list_resumes
    from src.tracker import excel
    from src.llm.factory import get_client
    from src.run_core import execute as run_core_execute, search_titles
    from src.store.runs import fail_run, start_run

    llm = get_client(config)
    print(f"  LLM: {llm.provider} / {llm.model}")

    output_dir = _BASE / config.get("output_dir", "output/")
    output_dir.mkdir(exist_ok=True)

    tracker_path = Path(config.get("tracker_path", "../job_tracker.xlsx"))
    if not tracker_path.is_absolute():
        tracker_path = (_BASE / tracker_path).resolve()

    apify_cfg     = config.get("apify", {})
    search_cfg    = config.get("search", {})
    days_posted   = search_cfg.get("days_posted", 7)
    jobs_per_cat  = search_cfg.get("jobs_per_category", 50)
    run_timeout   = search_cfg.get("run_timeout", 120)
    shortlist_min = config.get("tailoring", {}).get("shortlist_score", 7)

    print("=" * 62)
    print("  Job Hunt Automation")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 62)

    # ── 1. Read the profile ────────────────────────────────────────────────
    print("\n[1/7] Reading your profile...")
    conn, user_id = _open_store()
    try:
        resumes = list_resumes(conn, user_id, active_only=True)
        profile = get_profile(conn, user_id)
    finally:
        conn.close()

    if not resumes:
        print("  ERROR: no active resume on file, so there is nothing to "
              "search for.")
        print("  Run:  python main.py --serve")
        print("  Then open http://127.0.0.1:5000/setup and upload one.")
        sys.exit(1)

    titles = search_titles(resumes)
    if not titles:
        print("  ERROR: your resumes have no target roles between them.")
        print("  Open http://127.0.0.1:5000/profile and add at least one.")
        sys.exit(1)

    work_modes = profile["work_modes"] or ["remote"]
    print(f"  {len(resumes)} resume(s): "
          f"{', '.join(r['label'] for r in resumes)}")
    print(f"  Searching: {', '.join(titles)}")
    print(f"  Where: {profile['location'] or 'remote'} "
          f"({', '.join(work_modes)})")

    # ── 2. Scrape (steps 2-5 happen inside run_core.execute) ────────────────
    if not os.environ.get("APIFY_TOKEN"):
        print("\n  ERROR: APIFY_TOKEN not set.")
        print("  PowerShell:  $env:APIFY_TOKEN = 'apify_api_xxx'")
        sys.exit(1)

    print(f"\n[2/7] Scraping Indeed + LinkedIn + Remote boards "
          f"(last {days_posted} days)...")

    # Company enrichment is disabled by default; the block below — unchanged
    # from before this refactor — only runs when config turns it on. It sits
    # inside run_core's enrich hook because that is the one point between
    # dedupe and scoring that both callers of execute() share.
    enrichment_enabled = config.get("company_enrichment", {}).get("enabled", False)

    def _enrich(new_jobs: list[dict]) -> list[dict]:
        print("\n[3/7] Deduplicating...")
        print(f"  {len(new_jobs)} new jobs")

        if enrichment_enabled:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            from src.agents.company_agent import lookup as company_lookup
            from src.filters.company_filter import CompanyFilter
            from src.tracker.cache import CompanyCache

            unique_companies = list({j["company"] for j in new_jobs if j["company"]})
            print(f"\n[4/7] Company enrichment ({len(unique_companies)} companies)...")
            cache = CompanyCache(
                ttl_days=config.get("company_filter", {}).get("cache_ttl_days", 7))
            company_intel: dict[str, dict] = {}

            def _fetch_intel(company: str) -> tuple[str, dict]:
                return company, company_lookup(company, config, llm, cache)

            with ThreadPoolExecutor(max_workers=1) as pool:
                futures = {pool.submit(_fetch_intel, c): c for c in unique_companies}
                for done_idx, fut in enumerate(as_completed(futures), start=1):
                    c, intel = fut.result()
                    company_intel[c] = intel
                    if done_idx % 10 == 0:
                        print(f"  {done_idx}/{len(unique_companies)} resolved")
            print(f"  Done — {len(company_intel)} companies enriched")

            filt = CompanyFilter(config)
            filtered = 0
            for job in new_jobs:
                intel  = company_intel.get(job["company"],
                                           {"stage": "unknown", "growth_score": 5})
                result = filt.evaluate(job["company"], intel)
                job["_intel"]  = intel
                job["_filter"] = result
                if result["verdict"] == "REJECT":
                    filtered += 1
            print(f"  {len(new_jobs) - filtered} pass | {filtered} filtered")
        else:
            print("\n[4/7] Company enrichment: disabled "
                  "(set company_enrichment.enabled: true in config.yaml "
                  "to activate)")
            _passthrough = {"stage": "", "growth_score": "", "headcount_range": "",
                            "source": "", "verdict": "PASS"}
            for job in new_jobs:
                job["_intel"]  = _passthrough
                job["_filter"] = {"verdict": "PASS", "reason": "", "warn": None}

        print(f"\n[5/7] Scoring {len(new_jobs)} jobs against "
              f"{len(resumes)} resume(s)...")
        return new_jobs

    printed_steps: set[tuple[str, str, str]] = set()

    def _announce(event: dict) -> None:
        """Print each scrape step as it finishes, as this always has."""
        for step in event["steps"]:
            _announce_step(step, printed_steps)

    conn, user_id = _open_store()
    try:
        run_id = start_run(conn, user_id)
        try:
            result = run_core_execute(
                conn, user_id, run_id, config, resumes, profile,
                on_progress=_announce, enrich=_enrich)
        except Exception as exc:
            fail_run(conn, run_id, str(exc))
            raise
    finally:
        conn.close()

    print(f"\n  Total scraped: {result.scraped}")

    # ── Build the tracker rows from what the core scored ───────────────────
    cat_order = {title: i for i, title in enumerate(result.titles)}
    by_resume = {r["id"]: r for r in resumes}
    rows: list[dict] = []
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    for job in result.scored_jobs:
        intel  = job.get("_intel", {})
        filt_r = job.get("_filter", {"verdict": "PASS", "reason": "",
                                     "warn": None})
        is_filtered = filt_r["verdict"] == "REJECT"

        row: dict = {
            # Columns 1-15 — same positions as original job_tracker_updater.py
            "Search Category":    job["cat"],
            "Job Title":          job["title"],
            "Company":            job["company"],
            "Location":           job["location"],
            "Platform":           job["platform"],
            "Posting Date":       job["date"],
            "Apply Link":         job["url"],
            "Match Score":        job["match_score"],
            # Removed — was a hardcoded lookup table shown as a percentage.
            # Column kept blank so existing tracker files stay aligned.
            "Interview Chance":   "",
            "Recommended Resume": job["resume_label"],
            "Tailoring Needed":   job["tailoring"],
            "Application Status": "Filtered Out" if is_filtered else "New",
            "Date Applied":       "",
            "Follow-up Date":     "",
            "Notes":              filt_r.get("warn", "") or "",
            # Columns 16+ — new fields
            "Salary":             job.get("salary", ""),
            "Apply_Now":          "",
            "LLM_Reason":         filt_r.get("reason", "") if is_filtered else "",
            "Risk_Flags":         "",
            "Company Size":       intel.get("headcount_range", ""),
            "Company Stage":      intel.get("stage", ""),
            "Growth Score":       intel.get("growth_score", ""),
            "Intel Source":       intel.get("source", ""),
            "Loaded At":          run_ts,
            # Internal — not written to Excel
            "_score":             job["match_score"],
            "_description":       job.get("description", ""),
            "_cat_order":         cat_order.get(job["cat"], 9),
            "_resume_id":         job["resume_id"],
            "Low Growth Signal":  bool(filt_r.get("warn")),
        }
        rows.append(row)

    rows.sort(key=lambda r: (r["_cat_order"], -r["_score"]))
    filtered_out_count = sum(1 for r in rows
                             if r["Application Status"] == "Filtered Out")

    # ── 7. LLM shortlist decisions ─────────────────────────────────────────
    print(f"\n[6/7] LLM shortlist decisions (score >= {shortlist_min})...")
    shortlist = [r for r in rows
                 if r["_score"] >= shortlist_min
                 and r["Application Status"] != "Filtered Out"]
    print(f"  {len(shortlist)} jobs in shortlist")

    for row in shortlist:
        winner = by_resume.get(row["_resume_id"]) or resumes[0]
        variant_file = output_dir / f"resume_{_safe(row['Search Category'])}.txt"
        if variant_file.exists():
            resume_text = variant_file.read_text(encoding="utf-8")
        else:
            resume_text = winner.get("extracted_text", "")

        # Fall back to the title line only when the actor returned no body.
        jd_text = row["_description"] or f"{row['Job Title']} at {row['Company']}"
        decision = shortlist_decision(
            job_description=jd_text,
            resume_text=resume_text,
            role_title=row["Job Title"],
            company_name=row["Company"],
            match_score=row["_score"],
            llm=llm,
        )
        row["Apply_Now"]   = decision.get("apply_now", "")
        row["LLM_Reason"]  = decision.get("reason", "")
        row["Risk_Flags"]  = decision.get("risk_flags", "")

    # Pre-generate role variants for apply_now jobs
    apply_now_jobs = [r for r in rows if r.get("Apply_Now") == "yes"]
    if apply_now_jobs:
        print(f"\n  Generating resume variants for {len(apply_now_jobs)} apply-now jobs...")
        # rows is already sorted by (cat_order, -_score), so the first apply-now
        # row seen for a category is that category's highest-scoring job — use
        # its winning resume, not an arbitrary one.
        role_resume: dict[str, dict] = {}
        for r in apply_now_jobs:
            role_resume.setdefault(r["Search Category"],
                                   by_resume.get(r["_resume_id"]) or resumes[0])
        for role, winner in role_resume.items():
            try:
                generate_role_variant(role, winner, output_dir, llm)
            except Exception as e:
                logger.warning("Role variant failed for %s: %s", role, e)

    # ── Save ───────────────────────────────────────────────────────────────
    print(f"\n[7/7] Saving tracker...")
    if tracker_path.exists() and tracker_path.stat().st_size == 0:
        tracker_path.unlink()
    try:
        excel.save(tracker_path, rows)
    except PermissionError:
        print(f"\n  ERROR: Close {tracker_path.name} in Excel and retry.")
        sys.exit(1)

    # ── Summary ────────────────────────────────────────────────────────────
    high = [r for r in rows if r["_score"] >= 8
            and r["Application Status"] != "Filtered Out"]
    apply_now = [r for r in rows if r.get("Apply_Now") == "yes"]
    by_cat: dict[str, int] = {}
    by_platform: dict[str, int] = {}
    for r in rows:
        by_cat[r["Search Category"]] = by_cat.get(r["Search Category"], 0) + 1
        by_platform[r["Platform"]] = by_platform.get(r["Platform"], 0) + 1

    print("\n" + "=" * 62)
    print(f"  Done!  {len(rows)} new jobs added.")
    print(f"  Filtered out: {filtered_out_count}  |  Shortlisted: {len(shortlist)}")
    print(f"  Apply_Now: {len(apply_now)}")
    if by_cat:
        platform_summary = " ".join(f"{name}={count}" for name, count
                                    in sorted(by_platform.items()))
        print(f"\n  By platform: {platform_summary}")
        print(f"  By category:")
        for cat, n in sorted(by_cat.items(), key=lambda x: cat_order.get(x[0], 9)):
            print(f"    {cat:28s}: {n}")
    if high:
        print(f"\n  * {len(high)} high-scoring jobs (>=8):")
        for r in high[:5]:
            flag = " → APPLY NOW" if r.get("Apply_Now") == "yes" else ""
            print(f"    [{r['_score']}/10] {r['Job Title']} @ {r['Company']}{flag}")
    print(f"\n  Tracker: {tracker_path}")
    print("=" * 62)

    print(f"  Stored {result.kept} roles.")


def _serve(config: dict, port: int) -> None:
    """Run the local app. Binds to localhost only."""
    import webbrowser

    from src.app import create_app

    shortlist_min = config.get("tailoring", {}).get("shortlist_score", 7)
    app = create_app(user_name=DEFAULT_USER, shortlist_min=shortlist_min)
    url = f"http://127.0.0.1:{port}/"
    print(f"\n  Opportunity map: {url}")
    print(f"  Your resumes:    http://127.0.0.1:{port}/profile")
    print("  Ctrl-C to stop.")
    webbrowser.open(url)
    app.run(host="127.0.0.1", port=port, debug=False)


# ── Tailor one job ─────────────────────────────────────────────────────────────
def run_tailor(config: dict, jd_file: str) -> None:
    from src.agents.resume_agent import _safe
    from src.agents.tailoring_agent import tailor_job, save_tailoring_output
    from src.store.profile import list_resumes
    from src.tracker import excel
    from src.llm.factory import get_client
    from pathlib import Path

    llm = get_client(config)
    print(f"  LLM: {llm.provider} / {llm.model}")

    output_dir = _BASE / config.get("output_dir", "output/")
    output_dir.mkdir(exist_ok=True)
    tracker_path = Path(config.get("tracker_path", "../job_tracker.xlsx"))
    if not tracker_path.is_absolute():
        tracker_path = (_BASE / tracker_path).resolve()

    jd_path = Path(jd_file)
    if not jd_path.exists():
        print(f"ERROR: JD file not found: {jd_path}")
        sys.exit(1)
    jd_text = jd_path.read_text(encoding="utf-8")

    job_url = input("Enter job URL to look up in tracker (or press Enter to skip): ").strip()
    existing = excel.read_existing(tracker_path)

    row_data: dict = {}
    if job_url and job_url in existing:
        row_data = existing[job_url]
        print(f"  Found: {row_data.get('Job Title')} @ {row_data.get('Company')}")
    else:
        if job_url:
            print("  URL not found in tracker.")
        company_name = input("Company name: ").strip()
        role_title   = input("Role title: ").strip()
        row_data = {"Company": company_name, "Job Title": role_title,
                    "Search Category": ""}

    from src.store.profile import list_resumes

    conn, user_id = _open_store()
    try:
        active = list_resumes(conn, user_id, active_only=True)
    finally:
        conn.close()
    if not active:
        print("ERROR: no active resume on file. Upload one at /setup first.")
        sys.exit(1)
    resume = active[0]

    # A cached role variant is a better starting point than the raw resume,
    # because it was already rewritten for this kind of role.
    category  = row_data.get("Search Category", "")
    safe_role = _safe(category)
    variant_file = output_dir / f"resume_{safe_role}.txt" if safe_role else None
    if variant_file and variant_file.exists():
        resume_text = variant_file.read_text(encoding="utf-8")
    else:
        resume_text = resume.get("extracted_text", "")

    result = tailor_job(
        jd_text, resume_text,
        row_data.get("Job Title", ""),
        row_data.get("Company", ""),
        llm,
        resume=resume,
    )
    if result is None:
        print("ERROR: Tailoring failed. Check your LLM API key in job_hunt/.env.")
        sys.exit(1)

    folder = save_tailoring_output(result, row_data.get("Company", "company"), output_dir)
    print(f"\n  Ready → {folder}/")
    print(f"  cover_letter.txt + resume_highlights.txt")

    # Update tracker status
    if job_url and job_url in existing:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(tracker_path)
            ws = wb.active
            url_col = None
            status_col = None
            for ci in range(1, ws.max_column + 1):
                h = ws.cell(1, ci).value
                if h == "Apply Link":
                    url_col = ci
                if h == "Application Status":
                    status_col = ci
            if url_col and status_col:
                for ri in range(2, ws.max_row + 1):
                    cell = ws.cell(ri, url_col)
                    if (cell.hyperlink and cell.hyperlink.target == job_url):
                        ws.cell(ri, status_col).value = "Tailored"
                        break
            wb.save(tracker_path)
            print("  Tracker updated → Tailored")
        except PermissionError:
            print(f"  Warning: Close {tracker_path.name} to update status.")


# ── Entry ──────────────────────────────────────────────────────────────────────
def main() -> None:
    config = _load_config()
    _check_deps(config)
    parser = argparse.ArgumentParser(description="Job Hunt Automation")
    parser.add_argument("--tailor", action="store_true",
                        help="Tailor one job from a JD file")
    parser.add_argument("--jd", type=str, default="",
                        help="Path to job description .txt file (required with --tailor)")
    parser.add_argument("--serve", action="store_true",
                        help="Serve the app without scraping")
    parser.add_argument("--port", type=int, default=5000,
                        help="Port for the local app (default 5000)")
    args = parser.parse_args()

    if args.tailor:
        if not args.jd:
            print("ERROR: --jd <file.txt> is required with --tailor")
            sys.exit(1)
        run_tailor(config, args.jd)
    elif args.serve:
        _serve(config, args.port)
    else:
        run_pipeline(config)
        _serve(config, args.port)


if __name__ == "__main__":
    main()
