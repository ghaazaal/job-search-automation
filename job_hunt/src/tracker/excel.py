"""Excel tracker — read/write with all columns including new company + LLM columns."""
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

NAVY = "1E3A5F"
WHITE = "FFFFFF"
SCORE_HIGH = "D1FAE5"
SCORE_MED  = "FEF3C7"
SCORE_LOW  = "FEE2E2"
LOW_GROWTH = "FEF9C3"
FILTERED_BG = "F3F4F6"

CAT_COLORS = {
    "Analytics Engineer": "DBEAFE",
    "Product Analyst":    "DCFCE7",
    "Data Engineer":      "FEF9C3",
    "Data Analyst":       "FCE7F3",
}
STATUS_COLORS = {
    "New":            "EFF6FF",
    "Ready to Apply": "FEF9C3",
    "Tailored":       "E0F2FE",
    "Applied":        "DCFCE7",
    "Interview":      "D8B4FE",
    "Rejected":       "FEE2E2",
    "Ignored":        "F3F4F6",
    "Filtered Out":   "F3F4F6",
}

THIN  = Side(style="thin",   color="D1D5DB")
THICK = Side(style="medium", color="9CA3AF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Single source of truth for column definitions.
# Add new columns here — save() and _build_header() derive everything from this list.
COLS = [
    # ── Core job data ───────────────────────────────────────
    ("Search Category",    16),
    ("Job Title",          42),
    ("Company",            24),
    ("Location",           16),
    ("Salary",             18),
    ("Platform",           10),
    ("Posting Date",       13),
    ("Apply Link",         12),
    # ── Scoring ─────────────────────────────────────────────
    ("Match Score",        13),
    ("Interview Chance",   16),
    ("Tailoring Needed",   16),
    # ── Status & workflow ───────────────────────────────────
    ("Application Status", 18),
    ("Apply_Now",          10),
    ("LLM_Reason",         40),
    ("Risk_Flags",         30),
    ("Date Applied",       13),
    ("Follow-up Date",     13),
    ("Notes",              40),
    # ── Company enrichment (populated when enabled) ─────────
    ("Company Size",       14),
    ("Company Stage",      14),
    ("Growth Score",       12),
    ("Intel Source",       12),
    # ── Meta ────────────────────────────────────────────────
    ("Loaded At",          18),
]
COL_NAMES = [c[0] for c in COLS]
CAT_ORDER = {"Analytics Engineer": 0, "Data Engineer": 1,
             "Data Analyst": 2, "Product Analyst": 3}


def read_existing(path: Path) -> dict[str, dict]:
    """Return {url: row_data} for all rows in the tracker."""
    existing: dict[str, dict] = {}
    if not path.exists():
        return existing
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=False):
            url = None
            row_data: dict = {}
            for cell in row:
                h = ws.cell(1, cell.column).value
                if not h:
                    continue
                row_data[h] = cell.value
                if cell.hyperlink:
                    url = cell.hyperlink.target
            if url:
                status = (row_data.get("Application Status")
                          or row_data.get("Status") or "New")
                if status == "Not Applied":
                    status = "New"
                row_data["Application Status"] = status
                existing[url] = row_data
    except Exception as e:
        print(f"  Warning: could not fully read existing tracker: {e}")
    return existing


def _style_row(ws, row_idx: int, row_data: dict) -> None:
    cat    = row_data.get("Search Category", "")
    status = row_data.get("Application Status", "New")
    is_filtered   = status == "Filtered Out"
    is_low_growth = row_data.get("Low Growth Signal", False)

    bg = (FILTERED_BG if is_filtered
          else STATUS_COLORS.get(status, CAT_COLORS.get(cat, "FFFFFF")))

    for ci, col_name in enumerate(COL_NAMES, start=1):
        val  = row_data.get(col_name, "")
        cell = ws.cell(row_idx, ci, val)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = BORDER
        cell.alignment = Alignment(
            vertical="center",
            horizontal="left" if ci in {2, 3, 14, 15, 18} else "center",
            wrap_text=(ci in {2, 14, 15, 18}),
        )
        cell.font = Font(
            size=10, name="Calibri",
            color="9CA3AF" if is_filtered else "000000",
        )

        if col_name == "Apply Link" and val:
            cell.value     = "Apply →"
            cell.hyperlink = val
            cell.font      = Font(color="1D4ED8", underline="single",
                                  size=10, name="Calibri")
        elif col_name == "Match Score" and isinstance(val, int):
            if val >= 7:
                cell.fill = PatternFill("solid", fgColor=SCORE_HIGH)
                cell.font = Font(bold=True, color="065F46", size=10, name="Calibri")
            elif val >= 5:
                cell.fill = PatternFill("solid", fgColor=SCORE_MED)
                cell.font = Font(bold=True, color="92400E", size=10, name="Calibri")
            else:
                cell.fill = PatternFill("solid", fgColor=SCORE_LOW)
                cell.font = Font(color="991B1B", size=10, name="Calibri")
        elif col_name == "Growth Score" and is_low_growth and not is_filtered:
            cell.fill = PatternFill("solid", fgColor=LOW_GROWTH)
        elif col_name == "Search Category":
            cell.fill = PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FFFFFF"))
            cell.font = Font(bold=True, size=9, name="Calibri", color=NAVY)
        elif col_name == "Application Status":
            cell.fill = PatternFill("solid",
                                    fgColor=STATUS_COLORS.get(str(val), "F9FAFB"))
            cell.font = Font(bold=True, size=9, name="Calibri")
        elif col_name == "Apply_Now" and val == "yes":
            cell.font = Font(bold=True, color="065F46", size=10, name="Calibri")

    ws.row_dimensions[row_idx].height = 20


def _build_header(ws) -> None:
    """Write (or overwrite) row 1 with the canonical COLS header."""
    for ci, (name, width) in enumerate(COLS, start=1):
        cell = ws.cell(1, ci, name)
        cell.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        l = THICK if ci == 1 else THIN
        r = THICK if ci == len(COLS) else THIN
        cell.border = Border(left=l, right=r, top=THICK, bottom=THICK)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def _drop_blank_trailing_cols(ws) -> None:
    """Delete completely empty columns beyond our schema width."""
    while ws.max_column > len(COLS):
        ci = ws.max_column
        if all(ws.cell(r, ci).value is None for r in range(1, ws.max_row + 1)):
            ws.delete_cols(ci)
        else:
            break


def save(path: Path, new_rows: list[dict]) -> None:
    """Append new_rows to the tracker.

    Always rebuilds the header row so column names are correct even when
    appending to a file created by an older version of this script.
    Drops blank trailing columns beyond our schema width.
    """
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active

        # Rebuild header — handles schema changes and missing column names
        _build_header(ws)
        _drop_blank_trailing_cols(ws)

        last_row = ws.max_row
        while last_row > 1 and all(
            ws.cell(last_row, c).value is None
            for c in range(1, len(COLS) + 1)
        ):
            last_row -= 1

        for row_data in new_rows:
            last_row += 1
            _style_row(ws, last_row, row_data)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Tracker"
        _build_header(ws)
        for ri, row_data in enumerate(new_rows, start=2):
            _style_row(ws, ri, row_data)

    wb.save(path)
