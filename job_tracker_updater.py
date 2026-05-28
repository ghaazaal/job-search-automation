"""
job_tracker_updater.py — Ghazal Izadi's Job Search ATS Updater
================================================================
Run this script whenever you want to refresh your job tracker with new listings.

HOW IT WORKS:
  1. Reads the existing tracker (job_tracker_ghazal.xlsx) — preserves all
     your statuses, notes, dates exactly as you left them.
  2. Runs Apify scrapers for all 4 role categories on BOTH Indeed AND LinkedIn.
  3. Compares incoming jobs against existing URLs — skips duplicates.
  4. Scores new jobs and appends them to the tracker.
  5. Saves the updated file.

SETUP:
  Set your Apify API token as an environment variable:
    Windows (PowerShell):  $env:APIFY_TOKEN = "apify_api_xxxxxxxx"
    Windows (cmd):         set APIFY_TOKEN=apify_api_xxxxxxxx
    Or paste it directly into APIFY_TOKEN below (not for shared machines).

  Install dependencies if needed:
    pip install openpyxl requests

USAGE:
  python job_tracker_updater.py

  Edit the CONFIG section below to change search parameters.
"""

import os
import re
import sys
import time
import json
import requests
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date

# =============================================================================
# CONFIG — edit these as needed
# =============================================================================
APIFY_TOKEN   = os.environ.get("APIFY_TOKEN", "")   # or paste token string here
TRACKER_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "job_tracker_ghazal.xlsx")
DAYS_POSTED   = 7     # fetch jobs from last N days (1, 3, 7, or 14)
JOBS_PER_CAT  = 50    # max jobs per category per platform per run
RUN_TIMEOUT   = 120   # seconds to wait for each Apify actor run

# Apify actor IDs
INDEED_ACTOR   = "valig~indeed-jobs-scraper"
LINKEDIN_ACTOR = "valig~linkedin-jobs-scraper"
APIFY_BASE     = "https://api.apify.com/v2/acts/{actor}/run-sync-get-dataset-items"

# Search categories — runs on BOTH Indeed and LinkedIn
CATEGORIES = [
    {"category": "Analytics Engineer", "title": "Analytics Engineer"},
    {"category": "Data Engineer",      "title": "Data Engineer"},
    {"category": "Data Analyst",       "title": "Data Analyst"},
    {"category": "Product Analyst",    "title": "Product Analyst"},
]

# =============================================================================
# CANDIDATE PROFILE
# =============================================================================
RESUMES = {
    "Analytics Engineer": "cv_ghazal_izadi_AE.pdf",
    "Data Engineer":      "Zahra (Ghazal) Izadi-Data Engineer.pdf",
    "Data Analyst":       "Zahra (Ghazal) Izadi-data analyst.pdf",
    "Product Analyst":    "Zahra (Ghazal) Izadi-product analyst.pdf",
}

# =============================================================================
# SCORING ENGINE
# =============================================================================
def score_job(title: str, category: str, salary_str: str, url: str) -> dict:
    t     = title.lower()
    url_l = url.lower()

    senior_kws = ["senior", "sr.", "sr ", "staff", "principal", "lead",
                  "manager", "director", "head of", "vp ", "vice president"]
    junior_kws = ["junior", "jr.", "jr ", "intern", "entry", "associate ",
                  "analyst i ", "engineer i "]
    exec_kws   = ["vp ", "vice president", "director", "head of", "executive"]

    is_senior = any(s in t for s in senior_kws)
    is_junior = any(j in t for j in junior_kws)
    is_exec   = any(e in t for e in exec_kws)

    core_ae  = ["analytics engineer", "analytical engineer", "analytics eng"]
    core_de  = ["data engineer", "data infrastructure", "data platform engineer",
                "data integration engineer", "etl engineer", "elt engineer",
                "data intelligence engineer"]
    core_da  = ["data analyst", "bi analyst", "business intelligence analyst",
                "bi engineer", "business intelligence engineer",
                "data systems analyst", "data cloud", "data mining"]
    core_pa  = ["product analyst", "product data analyst", "analytics product",
                "senior data analyst, product"]
    adjacent = ["data scientist", "analytics manager", "marketing analytics",
                "reporting analyst", "data content", "data strategy",
                "experimentation platform", "data & ai", "data solutions",
                "data architect", "database engineer", "cloud data engineer",
                "databricks", "data quality", "data governance"]
    weak     = ["software engineer", "devops", "backend engineer", "qa engineer",
                "field service", "telecom", "mobile engineer", "sales",
                "security engineer", "marketing manager", "hr analyst",
                "total rewards", "billing", "actuarial",
                "business analyst (claims)", "configuration", "paid search"]

    if   any(k in t for k in core_ae):   title_score = 40
    elif any(k in t for k in core_de):   title_score = 35
    elif any(k in t for k in core_da):   title_score = 35
    elif any(k in t for k in core_pa):   title_score = 30
    elif any(k in t for k in adjacent):  title_score = 20
    elif any(k in t for k in weak):      title_score = 5
    else:                                title_score = 15

    cat_bonus = {"Analytics Engineer": 10, "Data Engineer": 8,
                 "Data Analyst": 7, "Product Analyst": 5}.get(category, 5)

    if   is_exec:   seniority_score = 5
    elif is_junior: seniority_score = 8
    elif is_senior: seniority_score = 20
    else:           seniority_score = 14

    skill_kws = {
        "airflow": 4, "etl": 3, "elt": 3, "sql": 3, "power bi": 4,
        "clickhouse": 4, "python": 3, "dbt": 3, "data model": 4,
        "kpi": 3, "dashboard": 3, "pipeline": 3, "warehouse": 3,
        "medallion": 4, "quality": 2, "governance": 3, "analytics": 2,
        "bi": 2, "tableau": 2, "plotly": 3, "streamlit": 3,
        "experimentation": 3,
    }
    skill_score = 0
    for kw, pts in skill_kws.items():
        if kw in t or kw in url_l:
            skill_score = min(skill_score + pts, 20)

    penalty = 0
    if "w2 only"   in t or "w2 only"  in url_l: penalty += 15
    if "clearance" in t or "secret"   in t:      penalty += 30
    if "cpt"       in t or "opt"      in t:      penalty += 20
    if is_exec:                                   penalty += 10
    if any(k in t for k in ["coldfusion", "mainframe", "cobol", "sas "]):
        penalty += 25
    if any(k in t for k in ["mobile", "ios", "android", "front-end", "frontend",
                              "react", "vue", "angular", "ruby", "rails"]):
        penalty += 30

    raw         = max(0, min(title_score + cat_bonus + seniority_score
                             + skill_score - penalty, 90))
    match_score = max(1, round(raw / 9))

    base = {10: 45, 9: 45, 8: 45, 7: 30, 6: 30, 5: 15, 4: 15}.get(match_score, 5)
    sal_nums = re.findall(r'\d+', (salary_str or "").replace(',', ''))
    if sal_nums:
        top_sal = int(sal_nums[-1])
        if   top_sal >= 400_000: base -= 15
        elif top_sal >= 200_000: base -= 5
    interview_chance = f"{max(5, min(base, 60))}%"

    core_map = {"Analytics Engineer": core_ae, "Data Engineer": core_de,
                "Data Analyst": core_da, "Product Analyst": core_pa}
    if any(k in t for k in core_map.get(category, [])):
        tailoring = "Minor" if category != "Product Analyst" else "Moderate"
    elif match_score >= 6: tailoring = "Moderate"
    elif match_score >= 4: tailoring = "Significant"
    else:                  tailoring = "N/A"

    rec_resume = RESUMES.get(category, RESUMES["Analytics Engineer"])
    if "bi " in t or "business intelligence" in t:
        rec_resume = RESUMES["Analytics Engineer"]

    return {
        "match_score":      match_score,
        "interview_chance": interview_chance,
        "rec_resume":       rec_resume,
        "tailoring":        tailoring,
    }


# =============================================================================
# APIFY HELPERS
# =============================================================================
def _call_apify(actor_id: str, payload: dict, label: str) -> list:
    """POST to Apify run-sync endpoint, return raw items list."""
    if not APIFY_TOKEN:
        return []
    url    = APIFY_BASE.format(actor=actor_id)
    params = {"token": APIFY_TOKEN, "timeout": RUN_TIMEOUT}
    try:
        resp = requests.post(
            url, json=payload,
            headers={"Content-Type": "application/json"},
            params=params, timeout=RUN_TIMEOUT + 30
        )
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout:
        print(f"    ✗ Timeout for {label}")
        return []
    except requests.exceptions.RequestException as e:
        print(f"    ✗ Request error for {label}: {e}")
        return []
    except json.JSONDecodeError:
        print(f"    ✗ Could not parse response for {label}")
        return []


def _normalize_date(raw) -> str:
    today = date.today().isoformat()
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ",
                "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(raw)[:26], fmt).date().isoformat()
        except (ValueError, TypeError):
            pass
    return today


def _parse_salary(sal) -> str:
    if isinstance(sal, dict):
        v = sal.get("value") or {}
        if isinstance(v, dict):
            lo, hi = v.get("minValue"), v.get("maxValue")
            if lo and hi:
                return (f"${lo:,.0f} – ${hi:,.0f}"
                        if isinstance(lo, (int, float)) else f"{lo} – {hi}")
            return str(lo) if lo else ""
    return str(sal) if sal else ""


# =============================================================================
# INDEED SCRAPER  (country = "us", catches most remote-worldwide roles)
# =============================================================================
def scrape_indeed(category: str, title: str) -> list:
    date_map = {1: "last 24 hours", 3: "last 3 days",
                7: "last 7 days", 14: "last 14 days"}
    payload = {
        "title":      title,
        "location":   "remote",
        "country":    "us",
        "limit":      JOBS_PER_CAT,
        "datePosted": date_map.get(DAYS_POSTED, "last 7 days"),
    }
    label    = f"Indeed / {category}"
    raw      = _call_apify(INDEED_ACTOR, payload, label)
    today    = date.today().isoformat()
    jobs     = []

    for item in raw:
        url = (item.get("jobUrl") or item.get("applyUrl") or item.get("url") or "")
        if not url:
            continue
        jobs.append({
            "cat":      category,
            "title":    item.get("title") or item.get("jobTitle") or "",
            "company":  (item.get("employer") or {}).get("name") or item.get("company") or "",
            "location": item.get("location") or "Remote, US",
            "platform": "Indeed",
            "date":     _normalize_date(item.get("datePublished") or item.get("postedAt") or today),
            "url":      url,
            "salary":   _parse_salary(item.get("baseSalary") or item.get("salary") or {}),
        })
    return jobs


# =============================================================================
# LINKEDIN SCRAPER  (remote=["2"] → worldwide remote coverage)
# =============================================================================
# LinkedIn datePosted codes:
#   "r86400"   = last 24 hours
#   "r604800"  = last 7 days
#   "r2592000" = last 30 days
_LI_DATE_MAP = {
    1:  "r86400",
    7:  "r604800",
    14: "r604800",   # LinkedIn has no 14-day option; use 7-day
    30: "r2592000",
}

def scrape_linkedin(category: str, title: str) -> list:
    payload = {
        "title":       title,
        "location":    "Worldwide",
        "remote":      ["2"],          # "2" = Remote
        "datePosted":  _LI_DATE_MAP.get(DAYS_POSTED, "r604800"),
        "limit":       JOBS_PER_CAT,
    }
    label = f"LinkedIn / {category}"
    raw   = _call_apify(LINKEDIN_ACTOR, payload, label)
    today = date.today().isoformat()
    jobs  = []

    for item in raw:
        # LinkedIn actor field names vary — handle both common shapes
        url = (item.get("jobUrl") or item.get("applyUrl") or
               item.get("url") or item.get("link") or "")
        if not url:
            continue

        job_title = (item.get("title") or item.get("jobTitle") or
                     item.get("positionName") or "")

        # Company: may be nested or flat
        company_raw = item.get("company") or item.get("companyName") or ""
        if isinstance(company_raw, dict):
            company = company_raw.get("name") or ""
        else:
            company = str(company_raw)

        location = (item.get("location") or item.get("jobLocation") or "Remote")

        posted_raw = (item.get("datePublished") or item.get("postedAt") or
                      item.get("publishedAt") or item.get("listedAt") or today)

        jobs.append({
            "cat":      category,
            "title":    job_title,
            "company":  company,
            "location": location,
            "platform": "LinkedIn",
            "date":     _normalize_date(posted_raw),
            "url":      url,
            "salary":   _parse_salary(item.get("salary") or item.get("baseSalary") or {}),
        })
    return jobs


# =============================================================================
# EXCEL STYLES
# =============================================================================
NAVY       = "1E3A5F"
WHITE      = "FFFFFF"
CAT_COLORS = {
    "Analytics Engineer": "DBEAFE",
    "Product Analyst":    "DCFCE7",
    "Data Engineer":      "FEF9C3",
    "Data Analyst":       "FCE7F3",
}
STATUS_COLORS = {
    "New":            "EFF6FF",
    "Ready to Apply": "FEF9C3",
    "Tailored":       "E0F2FE",
    "Applied":        "DCFCE7",
    "Interview":      "D8B4FE",
    "Rejected":       "FEE2E2",
    "Ignored":        "F3F4F6",
}
SCORE_HIGH = "D1FAE5"
SCORE_MED  = "FEF3C7"
SCORE_LOW  = "FEE2E2"
THIN       = Side(style="thin", color="D1D5DB")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("Search Category",    16),
    ("Job Title",          42),
    ("Company",            24),
    ("Location",           16),
    ("Platform",           10),
    ("Posting Date",       13),
    ("Apply Link",         12),
    ("Match Score",        13),
    ("Interview Chance",   16),
    ("Recommended Resume", 32),
    ("Tailoring Needed",   16),
    ("Application Status", 18),
    ("Date Applied",       13),
    ("Follow-up Date",     13),
    ("Notes",              40),
]
COL_NAMES = [c[0] for c in COLS]
CAT_ORDER = {"Analytics Engineer": 0, "Data Engineer": 1,
             "Data Analyst": 2, "Product Analyst": 3}


def style_data_row(ws, row_idx, row_data):
    cat    = row_data.get("Search Category", "")
    status = row_data.get("Application Status", "New")
    bg     = STATUS_COLORS.get(status, CAT_COLORS.get(cat, "FFFFFF"))

    for ci, col_name in enumerate(COL_NAMES, start=1):
        val  = row_data.get(col_name, "")
        cell = ws.cell(row_idx, ci, val)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = BORDER
        cell.alignment = Alignment(
            vertical="center",
            horizontal="left" if ci in [2, 3, 10, 15] else "center",
            wrap_text=(ci in [2, 10, 15])
        )
        cell.font = Font(size=10, name="Calibri")

        if col_name == "Apply Link" and val:
            cell.value     = "Apply →"
            cell.hyperlink = val
            cell.font      = Font(color="1D4ED8", underline="single",
                                  size=10, name="Calibri")
        elif col_name == "Match Score" and isinstance(val, int):
            if val >= 7:
                cell.fill = PatternFill("solid", fgColor=SCORE_HIGH)
                cell.font = Font(bold=True, color="065F46", size=10, name="Calibri")
            elif val >= 5:
                cell.fill = PatternFill("solid", fgColor=SCORE_MED)
                cell.font = Font(bold=True, color="92400E", size=10, name="Calibri")
            else:
                cell.fill = PatternFill("solid", fgColor=SCORE_LOW)
                cell.font = Font(color="991B1B", size=10, name="Calibri")
        elif col_name == "Search Category":
            cell.fill = PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FFFFFF"))
            cell.font = Font(bold=True, size=9, name="Calibri", color=NAVY)
        elif col_name == "Application Status":
            cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(val, "F9FAFB"))
            cell.font = Font(bold=True, size=9, name="Calibri")

    ws.row_dimensions[row_idx].height = 20


def build_header(ws):
    THICK = Side(style="medium", color="9CA3AF")
    for ci, (name, width) in enumerate(COLS, start=1):
        cell = ws.cell(1, ci, name)
        cell.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        l = THICK if ci == 1 else THIN
        r = THICK if ci == len(COLS) else THIN
        cell.border = Border(left=l, right=r, top=THICK, bottom=THICK)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


# =============================================================================
# READ EXISTING TRACKER
# =============================================================================
def read_existing_tracker(path):
    existing = {}
    if not os.path.exists(path):
        print("  No existing tracker found — will create fresh.")
        return existing
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=False):
            url      = None
            row_data = {}
            for cell in row:
                h = ws.cell(1, cell.column).value
                if not h:
                    continue
                row_data[h] = cell.value
                if cell.hyperlink:
                    url = cell.hyperlink.target
            if url:
                status = (row_data.get("Application Status") or
                          row_data.get("Status") or "New")
                if status == "Not Applied":
                    status = "New"
                row_data["Application Status"] = status
                existing[url] = row_data
        print(f"  Loaded {len(existing)} existing jobs.")
    except Exception as e:
        print(f"  Warning: could not fully read existing tracker: {e}")
    return existing


# =============================================================================
# MAIN
# =============================================================================
def run_update():
    print("=" * 62)
    print("  Job Tracker Updater — Ghazal Izadi")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 62)

    # ── Step 1: Read existing tracker ────────────────────────────────────
    print("\n[1/5] Reading existing tracker...")
    existing      = read_existing_tracker(TRACKER_PATH)
    existing_urls = set(existing.keys())

    # ── Step 2: Scrape Indeed + LinkedIn for all 4 categories ────────────
    print(f"\n[2/5] Scraping Indeed + LinkedIn (last {DAYS_POSTED} days)...")

    if not APIFY_TOKEN:
        print("\n  ERROR: APIFY_TOKEN not set.")
        print("  Set it first then re-run:")
        print("    PowerShell:  $env:APIFY_TOKEN = 'apify_api_xxx'")
        print("    cmd:         set APIFY_TOKEN=apify_api_xxx")
        sys.exit(1)

    all_scraped = []
    for cat in CATEGORIES:
        category, title = cat["category"], cat["title"]

        # Indeed
        print(f"\n  Indeed / {category}...")
        indeed_jobs = scrape_indeed(category, title)
        print(f"    -> {len(indeed_jobs)} fetched")
        all_scraped.extend(indeed_jobs)
        time.sleep(1)

        # LinkedIn
        print(f"  LinkedIn / {category}...")
        li_jobs = scrape_linkedin(category, title)
        print(f"    -> {len(li_jobs)} fetched")
        all_scraped.extend(li_jobs)
        time.sleep(1)

    print(f"\n  Total scraped: {len(all_scraped)} jobs across both platforms")

    # ── Step 3: Deduplicate ───────────────────────────────────────────────
    print("\n[3/5] Deduplicating...")
    seen_batch = set()
    new_jobs   = []
    skip_dup = skip_exist = 0
    for job in all_scraped:
        url = job["url"]
        if url in seen_batch:
            skip_dup += 1; continue
        seen_batch.add(url)
        if url in existing_urls:
            skip_exist += 1; continue
        new_jobs.append(job)
    print(f"  {len(new_jobs)} new | {skip_exist} already tracked "
          f"| {skip_dup} batch duplicates")

    # ── Step 4: Score ─────────────────────────────────────────────────────
    print(f"\n[4/5] Scoring {len(new_jobs)} new jobs...")
    scored_new = []
    for job in new_jobs:
        s = score_job(job["title"], job["cat"], job["salary"], job["url"])
        scored_new.append({
            "Search Category":    job["cat"],
            "Job Title":          job["title"],
            "Company":            job["company"],
            "Location":           job["location"],
            "Platform":           job["platform"],
            "Posting Date":       job["date"],
            "Apply Link":         job["url"],
            "Match Score":        s["match_score"],
            "Interview Chance":   s["interview_chance"],
            "Recommended Resume": s["rec_resume"],
            "Tailoring Needed":   s["tailoring"],
            "Application Status": "New",
            "Date Applied":       "",
            "Follow-up Date":     "",
            "Notes":              "",
            "_score":             s["match_score"],
            "_cat_order":         CAT_ORDER.get(job["cat"], 9),
        })
    scored_new.sort(key=lambda r: (r["_cat_order"], -r["_score"]))

    # ── Step 5: Save ──────────────────────────────────────────────────────
    print(f"\n[5/5] Saving tracker...")

    if os.path.exists(TRACKER_PATH):
        wb = load_workbook(TRACKER_PATH)
        ws = wb.active
        last_row = ws.max_row
        while last_row > 1 and all(
            ws.cell(last_row, c).value is None for c in range(1, len(COLS) + 1)
        ):
            last_row -= 1
        for row_data in scored_new:
            last_row += 1
            style_data_row(ws, last_row, row_data)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Tracker"
        build_header(ws)
        for ri, row_data in enumerate(scored_new, start=2):
            style_data_row(ws, ri, row_data)

    wb.save(TRACKER_PATH)

    # ── Summary ───────────────────────────────────────────────────────────
    by_cat      = {}
    by_platform = {"Indeed": 0, "LinkedIn": 0}
    by_score    = {"high(8-10)": 0, "mid(5-7)": 0, "low(1-4)": 0}
    for r in scored_new:
        by_cat[r["Search Category"]] = by_cat.get(r["Search Category"], 0) + 1
        by_platform[r["Platform"]]   = by_platform.get(r["Platform"], 0) + 1
        s = r["_score"]
        if s >= 8:   by_score["high(8-10)"] += 1
        elif s >= 5: by_score["mid(5-7)"]   += 1
        else:        by_score["low(1-4)"]    += 1

    print("\n" + "=" * 62)
    print(f"  Done!  {len(scored_new)} new jobs added to tracker.")
    if scored_new:
        print(f"\n  By platform:  Indeed={by_platform.get('Indeed',0)}"
              f"  LinkedIn={by_platform.get('LinkedIn',0)}")
        print(f"\n  By category:")
        for cat, n in sorted(by_cat.items(), key=lambda x: CAT_ORDER.get(x[0], 9)):
            print(f"    {cat:28s}: {n}")
        print(f"\n  Score distribution: {by_score}")
    print(f"\n  Tracker: {TRACKER_PATH}")
    print("=" * 62)

    high = [r for r in scored_new if r["_score"] >= 8]
    if high:
        print(f"\n  * {len(high)} high-scoring new job(s) — consider tailoring:")
        for r in high[:5]:
            print(f"    [{r['_score']}/10] {r['Job Title']} @ {r['Company']}"
                  f"  ({r['Platform']})")
        if len(high) > 5:
            print(f"    ... and {len(high)-5} more in the tracker.")


if __name__ == "__main__":
    run_update()
